# pdf_tool.py
# PDFデータ抽出ツール（テキスト・表の両対応 / 一括処理対応）

import json
import sys
import os
import glob
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def load_config(config_path="config.json"):
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_text(pdf_path):
    results = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                for line in text.split("\n"):
                    line = line.strip()
                    if line:
                        results.append({"page": i + 1, "text": line})
    return results


def extract_tables(pdf_path):
    results = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            for table in tables:
                results.append({"page": i + 1, "table": table})
    return results


def get_header_style():
    font = Font(name="メイリオ", bold=True, color="FFFFFF", size=11)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    border = Side(style="thin", color="000000")
    border_style = Border(top=border, bottom=border, left=border, right=border)
    return font, fill, alignment, border_style


def get_data_style():
    font = Font(name="メイリオ", size=10)
    alignment = Alignment(vertical="center")
    border = Side(style="thin", color="000000")
    border_style = Border(top=border, bottom=border, left=border, right=border)
    return font, alignment, border_style


def apply_header_style(ws, row_num, col_count):
    h_font, h_fill, h_align, h_border = get_header_style()
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = h_border


def auto_adjust_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[column_letter].width = min(max_length + 4, 60)


def write_text_to_excel(text_data, output_path, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["ページ", "テキスト"])
    apply_header_style(ws, 1, 2)
    d_font, d_align, d_border = get_data_style()
    for item in text_data:
        ws.append([item["page"], item["text"]])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.font = d_font
            cell.alignment = d_align
            cell.border = d_border
    auto_adjust_width(ws)
    wb.save(output_path)


def write_tables_to_excel(table_data, output_path, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    current_row = 1
    for item in table_data:
        for row_idx, row in enumerate(item["table"]):
            ws.append(row)
            if row_idx == 0:
                apply_header_style(ws, current_row, len(row))
            current_row += 1
        ws.append([])
        current_row += 1
    d_font, d_align, d_border = get_data_style()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and not cell.font.bold:
                cell.font = d_font
                cell.alignment = d_align
                cell.border = d_border
    auto_adjust_width(ws)
    wb.save(output_path)


def process_single(pdf_path, output_path, mode, sheet_name):
    if not os.path.exists(pdf_path):
        print(f"エラー: {pdf_path} が見つかりません")
        return False
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) == 0:
                print(f"エラー: {pdf_path} にページがありません")
                return False
            print(f"  処理中: {pdf_path}（{len(pdf.pages)}ページ）")
    except Exception as e:
        print(f"  エラー: {pdf_path} を開けません → {e}")
        return False
    if mode == "text":
        data = extract_text(pdf_path)
        if not data:
            print("  スキップ: テキストが見つかりません")
            return False
        write_text_to_excel(data, output_path, sheet_name)
        print(f"  → {output_path}（{len(data)}行抽出）")
    elif mode == "table":
        data = extract_tables(pdf_path)
        if not data:
            print("  スキップ: テーブルが見つかりません")
            return False
        write_tables_to_excel(data, output_path, sheet_name)
        print(f"  → {output_path}（{len(data)}件抽出）")
    return True


def process_batch(input_folder, mode, sheet_name):
    if not os.path.isdir(input_folder):
        print(f"エラー: フォルダ {input_folder} が見つかりません")
        sys.exit(1)
    pdf_files = sorted(glob.glob(os.path.join(input_folder, "*.pdf")))
    if not pdf_files:
        print(f"エラー: {input_folder} 内にPDFファイルがありません")
        sys.exit(1)
    print(f"[一括処理] {len(pdf_files)}件のPDFを検出")
    print("=" * 50)
    success = 0
    fail = 0
    for pdf_path in pdf_files:
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        output_path = os.path.join(input_folder, f"{base_name}.xlsx")
        if process_single(pdf_path, output_path, mode, sheet_name):
            success += 1
        else:
            fail += 1
    print("=" * 50)
    print(f"[完了] 成功: {success}件 / 失敗: {fail}件 / 合計: {len(pdf_files)}件")


def main():
    config = load_config()
    mode = config["mode"]
    sheet_name = config["sheet_name"]
    batch = config.get("batch", False)
    if batch:
        process_batch(config["input_folder"], mode, sheet_name)
    else:
        pdf_path = sys.argv[1] if len(sys.argv) > 1 else config["input_pdf"]
        output_path = sys.argv[2] if len(sys.argv) > 2 else config["output_excel"]
        print(f"[モード] {'テキスト' if mode == 'text' else 'テーブル'}抽出")
        if process_single(pdf_path, output_path, mode, sheet_name):
            print("[完了]")
        else:
            sys.exit(1)


if __name__ == "__main__":
    main()